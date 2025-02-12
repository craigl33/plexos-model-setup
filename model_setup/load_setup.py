# -*- coding: utf-8 -*-
"""
Created on Mon Apr  8 18:58:31 2024

@author: HUNGERFORD_Z


## TODO

processing of the large frames (e.g. regional split of demand) might be better with dask?
legacy version allowed scaling the demand by end use, has not been implemented here
the code could be better modularised with something like the variables constructor from solution file processing
    -not sure if worth the effort to convert but for now multiple writes are together as they build on the same variables


"""


import pandas as pd
import numpy as np
import os
from model_setup.utils import add_pattern_index, make_pattern_index
from riselib.math import clean_timeseries
#from functions.read_weo import make_pattern_index, add_time_separators
from pyproj import CRS
from pathlib import Path
from openpyxl import load_workbook

class LoadSetup:
    def __init__(self, config):
        print("initialising class")
        self.c = config

        # sheet names frome the combined load sheet are read in to allow flexible identification
        if self.c.cfg['load']['load_type'] == 'WEO':
            self.excel_sheets = self._get_excel_sheet_names()
        else:
            self.excel_sheets = None

        self.demand_index = {}
        self.agg_demand_index = {}
        self.region_split = {}
        self.save_path = {}
        self.regions = self.c.get("parameters", "regions") 
        print("class initialised")


        self.gdf_adm = self.c.gdf_adm.copy()
        
        try:
           self.plot_save_path = Path(self.c.cfg['gis']['plot_save_path'])
           self.plot_save_path.mkdir(parents=True, exist_ok=True)
        except KeyError:
           self.plot_save_path = None

        self.default_crs = CRS('EPSG:4326')
        self.utm_crs = CRS('EPSG:3857')
        self.plot_col = self.c.cfg['transmission']['plot_col']
        self.modelling_reg_col = self.c.cfg['gis']['modelling_reg_col']
        
    def _get_excel_sheet_names(self):
        # This is all legacy stuff from the old code.
        # Basically, this is not needed anymore as the sheet names are passed directly to the class
        wb = load_workbook(None, read_only=True, data_only=False, keep_links=False)
        sheet_names = wb.sheetnames
        return(sheet_names)
    
    def create_demand_inputs_manual(self):
        """
        Alternative function for processing demand input from explicit demand inputs. 

        This function is built for use with non-WEO demand inputs.

        """



        
        
    def create_demand_inputs_WEO(self):
        
        # Directly retrieve the 'portfolio_assignments' section as a dictionary
        portfolio_assignments = self.c.get('portfolio_assignments')
        if portfolio_assignments is None:
            print("No [Portfolio assignments] entries not found in config file. Cannot run demand setup. Please complete at least one assignment for P1, etc.")
        
   
        for portfolio, settings in portfolio_assignments.items():
            
            # identify the relevant sheet names - flexible in ordering and to use scenario code or year alone
            # could have multi-match issues (returns first match) with the load sheet if you have other sheets with the scenario code or year that are not the index or factors
            load_sheet = self._find_sheet(self.excel_sheets, primary_id=settings['scenario_code'], fallback_id=settings['year'], exclusions=["DSM_index", "RegionalFactors"])
            index_sheet = self._find_sheet(self.excel_sheets, primary_id=settings['scenario_code'], fallback_id=settings['year'], keyword = "DSM_index")
            splitting_sheet = self._find_sheet(self.excel_sheets, primary_id=settings['scenario_code'], fallback_id=settings['year'], keyword = "RegionalFactors")
            
            # read in the hourly demand
            # LEGACY version used to have the capability to merge multiple regional sheets but this has not been integrated
            # hourly demand passed to the function to avoid clogging the memory
            hourly_demand = pd.read_excel(self.c.get('load', 'load_path'), load_sheet, skiprows=1)             
            
            # the index and regional split saved as class attributes as they are not big
            self.region_split[portfolio] = pd.read_excel(self.c.get('load', 'load_path'), splitting_sheet)
            
            # read the index and do some reformatting, done in the script to minimise manual adjustments
            demand_response_index = pd.read_excel(self.c.get('load', 'load_path'), index_sheet)
            di = demand_response_index[pd.notna(demand_response_index['Unnamed: 0'])]
            di = di[['Unnamed: 0', 'Faisability factor - shifting', 'Faisability factor - shedding',
                     'Technology x Acceptability factor', 'shed_header', 'shift_header', 'approx_unit', 'max_scale', 'max_shift']]
            di.columns = ['end_use', 'shift_factor', 'shed_factor', 'availability_factor', 'shed_header', 'shift_header', 'approx_unit', 'max_scale', 'max_shift']
            self.demand_index[portfolio] = di.copy()
            # note that this is not designed to maintain different values of max shift for the same end use
            self.agg_demand_index[portfolio] = di.groupby(['shift_header'])[["max_scale", "max_shift"]].max().reset_index()
            
            # calculate the regional split (returns database format)
            regional_demand_frame = self._make_regional_demand_split(hourly_demand, portfolio)
            # regional_demand_frame = regional_frame[["datetime", "end_use", "region", "load"]]
            
            # apply scaling factor
            
            if settings["load_scaling"] <= 2:                
                regional_demand_frame["load"] = regional_demand_frame["load"] * settings["load_scaling"]
            elif settings["load_scaling"] > 2:
                demand_total = regional_demand_frame["load"].sum()/1000000
                scaling_factor =  settings["load_scaling"] / demand_total
                regional_demand_frame["load"] = regional_demand_frame["load"] * scaling_factor
                
            print("Annual demand total after scaling equal to " + str(round(regional_demand_frame["load"].sum()/1000000, 1)) + " TWh")

            # combine with demand response index             
            indexed_demand = pd.merge(regional_demand_frame, self.demand_index[portfolio], how = "left")

            # add pattern index for input format
            indexed_demand = add_pattern_index(indexed_demand)            
            
            # ensure the save path is present for the portfolio
            self.save_path[portfolio] = self.c.get("path", "load_save_path") + settings["scenario_code"]
            # Check if the folder exists
            if not os.path.exists(self.save_path[portfolio]):
                # If the folder does not exist, create it
                os.makedirs(self.save_path[portfolio])
                print(f"The folder '{self.save_path[portfolio]}' was created.")
            else:
                print(f"The folder '{self.save_path[portfolio]}' already exists.")
            
            
            self._write_regional_totals(indexed_demand, portfolio)
            self._write_end_use_totals(indexed_demand, portfolio)
            self._write_shiftable_inputs(indexed_demand, portfolio)
            self._write_sheddable_inputs(indexed_demand, portfolio)
            
            # make totals
            
    def _find_sheet(self, sheet_names, primary_id, fallback_id, keyword=None, exclusions=None):
        """
        Find the sheet name that contains the primary_id or fallback_id, and optionally a keyword.

        This function is used to find the sheet name that contains the primary_id or fallback_id, and optionally a keyword.
        If a keyword is provided, the function will first try to find a sheet with the keyword. If no sheet is found, or no keyword is provided, the function will look for a sheet with the primary_id or fallback_id.
        
        This function appears to give incorrect results when there are multiple sheets with the keyword. It will return the first match, which may not be the correct sheet. This is a limitation of the function.
        
        """
    
        if exclusions is None:
            exclusions = []

        # If using the standard format of 'keyword_primary_id', check if there is an explicit match
        # Issues seem to occur with this function and it should be reviewed
        if keyword:
            explicit_match = f"{keyword}_{primary_id}"
        else:
            explicit_match = f"{primary_id}"

        if explicit_match in sheet_names:
            return explicit_match
        
        # To check if primary_id is a component of other primary_ids
        if len([(primary_id in i)&(keyword in i) for i in sheet_names] ) > 1:
            # Add the primary_ids which include the primary_id as a component to the exclusions list
            # This is to avoid returning a sheet with a primary_id that is a component of another primary_id
            # For example, if primary_id is 'P1' and there is a sheet with primary_id 'P1A', we don't want to return that sheet
            duplicate_ids = [i for i in sheet_names if (primary_id in i)&(keyword in i)]
            exclusions.extend([i for i in duplicate_ids if len(i) > min([len(j) for j in duplicate_ids])])

        # First, try to find a sheet with the keyword if it's provided
        # This is problematic if there are multiple sheets with the keyword, as it will return the first match
        if keyword:
            for sheet in sheet_names:
                if any(exclusion in sheet for exclusion in exclusions):
                    continue
                if (keyword in sheet) and (primary_id in sheet):
                    return sheet
                # If the keyword is not in the sheet name, but the fallback_id is, return the sheet
                elif (keyword in sheet) and (fallback_id in sheet) and (primary_id not in sheet):
                    print(f"Sheet with primary_id '{primary_id}' not found. Using sheet with fallback_id '{fallback_id}' for {keyword} keyword.")
                    return sheet

        # If no sheet found with the keyword, or no keyword provided, look for primary_id or fallback_id
        for sheet in sheet_names:
            if any(exclusion in sheet for exclusion in exclusions):
                continue
            if (primary_id in sheet):
                print(f"Sheet with primary_id '{primary_id}' found.")
                return sheet
            elif(fallback_id in sheet) and (primary_id not in sheet):
                print(f"Sheet with primary_id '{primary_id}' not found. Using sheet with fallback_id '{fallback_id}' for {keyword} keyword.")
                return sheet
    
        return None
    
    def _make_regional_demand_split(self, hourly_demand, portfolio):
        
        demand_melted = pd.melt(hourly_demand, id_vars='Unnamed: 0')
        demand_melted.columns = ['datetime', 'end_use', 'value']
        
        region_suffix = "_" + self.c.get("parameters", "model_region").upper()
        
        # Use apply() with a lambda function to conditionally remove the region suffix if present
        demand_melted['end_use'] = demand_melted['end_use'].apply(lambda x: x.replace(region_suffix, '') if region_suffix in x else x).str.upper()

        
        # drop entries not present in the index (empty columns or totals etc.)        
        demand_melted = demand_melted[demand_melted.end_use.isin(np.unique(self.demand_index[portfolio].end_use))]
        
        # melt the regional splits
        regions_factors_melted = pd.melt(self.region_split[portfolio][["end_use"] + self.regions],
                                         id_vars="end_use", value_vars = self.regions)
        
        regions_factors_melted.columns = ["end_use", "region", "region_factor"]
        
        #merge the two
        regional_frame = pd.merge(demand_melted, regions_factors_melted, how = "left")
        regional_frame["load"] = regional_frame["value"] * regional_frame["region_factor"]
        
        return(regional_frame[["datetime", "end_use", "region", "load"]])    

    def _write_regional_totals(self, indexed_demand, portfolio):
        
        #create total demands and write to .csv 
        totals = indexed_demand.groupby(['datetime', 'pattern', 'region'])['load'].sum().round(2).reset_index()
        totals_table = totals.pivot_table(values='load', index=['datetime', 'pattern'], columns='region').reset_index()
        totals_table.drop(columns=['datetime']).to_csv(self.save_path[portfolio] + '/total_load.csv', index=False)
       
    def _write_end_use_totals(self, indexed_demand, portfolio):

        #create whole region, by end-use demand for checks (not used in model) 
        end_use_demands = indexed_demand.groupby(['datetime', 'pattern', 'end_use'])[['load']].sum().reset_index()
        end_use_table = end_use_demands.pivot_table(values='load', index=['datetime', 'pattern'], columns='end_use').reset_index()
        end_use_table.to_csv(self.save_path[portfolio] + '/end_use_load_check.csv', index=False)
        
    def _write_shiftable_inputs(self, indexed_demand, portfolio):
        
        # subset indexed demand frame to entries with valid shift header
        df = indexed_demand.copy().dropna(subset = ['shift_header'])

        # multiply each demand type by sheddability to get the shiftable portion
        df['DSM_shift'] = df.load * df.shift_factor * df.availability_factor

        #create aggregated frame by shift_header and region for making the various shifting outputs
        #as well as maxes frame for each aggregated end use
        # create header aggregated frame
        aggregated_DSM = df.groupby(['datetime', 'pattern', 'region', 'shift_header'])[['DSM_shift']].sum().reset_index()
        # create header maxes frame for bids
        aggregated_DSM_maxes = aggregated_DSM.groupby(['region', 'shift_header'])[['DSM_shift']].max().reset_index()
        
        # create "native" load from totals minus shiftable demands
        totals = indexed_demand.groupby(['datetime', 'pattern', 'region'])['load'].sum().round(2).reset_index()
        # aggregate for subtraction native load
        shiftable_totals = aggregated_DSM.groupby(['datetime', 'pattern', 'region'])['DSM_shift'].sum().reset_index()
        
        ## merge with totals for subtraction and subtract
        native = pd.merge(totals, shiftable_totals, how='left')
        native.load = native.load - native.DSM_shift
        ## reshape and write to .csv
        native_table = (native.pivot_table(values='load', index=['datetime', 'pattern'], columns='region').reset_index().drop(columns=['datetime']))
        native_table.to_csv(self.save_path[portfolio] + '/native_load.csv', index=False)
        
        # create inputs for shiftable loads: load profile, daily sums, bid limits, max and min shifts, and annual limit for Aluminium 
        # load profile - based on aggregated frame (already aggregated to shift category and region), create header 
        aggregated_DSM['NAME'] = aggregated_DSM.region + '_' + aggregated_DSM.shift_header
        # np.unique(shiftable.header)
        ### reshape and write to .csv
        shiftable_table = (aggregated_DSM.pivot_table(values='DSM_shift', index=['datetime', 'pattern'], columns='NAME').reset_index().round(2))
        shiftable_table.drop(columns=['datetime']).to_csv(self.save_path[portfolio] + '/DSM_shift.csv', index=False)
        
        # daily sums 
        shift_sums = shiftable_table.set_index('datetime').drop(columns='pattern')
        shift_sums = shift_sums.resample('D').sum() / 1000  ## converted from MWh to GWh
        make_pattern_index(shift_sums.round(5)).to_csv(self.save_path[portfolio] + '/DSM_dayLimits.csv')
        
        # bids - scale DSM value by max scale
        indexed_aggregate = pd.merge(aggregated_DSM, self.agg_demand_index[portfolio], how='left')
        indexed_aggregate["bids"] = indexed_aggregate.DSM_shift * indexed_aggregate.max_scale
        ## find max value by region/shift category
        bids = indexed_aggregate.groupby(['NAME'])[['bids']].max().reset_index()
        ## add pattern column
        bids['pattern'] = 'M1-12'
        ## set column names and write to .csv
        bids = bids[['NAME', 'pattern', 'bids']]
        bids.columns = ['NAME', 'pattern', '1']
        bids.to_csv(self.save_path[portfolio] + '/DSM_bidQuantities.csv', index=False)
        
        # max and min shift - tried changing this to monthly but produced infeasibilities
        #agreed that annual max basis may be more reasonable anyway as demand response load is more coincident than unmanaged load
        #so annual max basis may already be conservative
        # np.unique(minmaxes.header)
        indexed_aggregate['1'] = indexed_aggregate.DSM_shift * indexed_aggregate.max_shift
        ## drop null values
        minmaxes = indexed_aggregate[pd.notnull(indexed_aggregate['1'])]
        minmaxes.loc[:,'NAME'] = minmaxes.region + '_MaxShift' + minmaxes.max_shift.astype(int).astype(str) + 'h'
        ## find maxima of max shift values
        minmaxes = minmaxes.groupby(['NAME'])['1'].max().reset_index()
        # chkframe = minmaxes.groupby(["NAME"])["DSM"].max().reset_index()
        minmaxes['pattern'] = 'M1-12'
        mins_frame = minmaxes.copy(deep=True)
        mins_frame['1'] = mins_frame['1'] * -1
        mins_frame.NAME = mins_frame.NAME.str.replace('Max', 'Min', case=True)
        # minmaxes.append(mins_frame)[["NAME", "pattern", "1"]].to_csv(save_path + "DSM_MaxShift.csv", index = False)

        combined_df = pd.concat([minmaxes, mins_frame])[['NAME', 'pattern', '1']]
        combined_df.to_csv(self.save_path[portfolio] + '/DSM_MaxShift.csv', index=False)
        
    def _write_sheddable_inputs(self, indexed_demand, portfolio):
        
        df = indexed_demand.copy().dropna(subset = ['shed_header'])
        # multiply each demand type by sheddability to get the shiftable portion
        df['DSM_shed'] = df.load * df.shed_factor * df.availability_factor

        #create aggregated frame by shift_header and region for making the various shifting outputs
        #as well as maxes frame for each aggregated end use
        # create header aggregated frame
        aggregated_DSM = df.groupby(['datetime', 'pattern', 'region', 'shed_header'])[['DSM_shed']].sum().reset_index()
        
        # legacy setting for CHina - not tested / set up for the new inputs
        # Aluminium annual limit
        if self.c.get("parameters", "has_aluminium") == True:
            alframe = aggregated_DSM[aggregated_DSM.shed_header == 'Al']
            alframe = alframe.groupby(['NAME'])[['DSM']].sum().reset_index()
            alframe['1'] = alframe.DSM / 1000
            alframe['pattern'] = 'M1-12'
            alframe[['NAME', 'pattern', '1']].to_csv(save_path / 'Al_AnnualLim.csv', index=False)

        # create shed loads: units, max capacity per type, and rating per type and unit 
        # units 
        # start with end use aggregates
        sheddable = aggregated_DSM.copy()
        # sheddable = sheddable[sheddable.region == "NER"]
        # sheddable = sheddable[sheddable.header == "Shed1h"]
        ## add unit size info
        agg_shed_index = self.demand_index[portfolio].groupby(['shed_header'])[["approx_unit"]].max().reset_index()
        
        sheddable = pd.merge(sheddable, agg_shed_index, how='left')
        
        sheddable['NAME'] = sheddable.region + '_' + sheddable.shed_header
        # get max values and calculate number of units
        units = sheddable.groupby(['NAME']).max().reset_index()
        units['1'] = units.DSM_shed / units.approx_unit
        units['1'] = units['1'].astype(float).round()
        # replace zeros with 1 if capacity is > 0
        units.loc[(units['1'] == 0) & (units['DSM_shed'] > 0), '1'] = 1
        # units['1']
        units.pattern = 'M1-12'
        units[['NAME', 'pattern', '1']].to_csv(self.save_path[portfolio] + '/shed_units.csv', index=False)
        # max capacity per type - calculate from DSM and units
        mask = units['1'] != 0
        # Perform the division only where the mask is True
        units.loc[mask, 'size'] = units.loc[mask, 'DSM_shed'] / units.loc[mask, '1']
        # Set 'size' to 0 where the mask is False
        units.loc[~mask, 'size'] = 0
        unit_size = units[['NAME', 'pattern', 'size']]
        unit_size.columns = ['NAME', 'pattern', '1']
        unit_size.to_csv(self.save_path[portfolio] + '/shed_max_cap.csv', index=False)

        # rating per type and unit by dividing through aggregated timeseries by number of units 
        sheddable_pu = pd.merge(sheddable, units[['NAME', '1']], how='left')
        mask = sheddable_pu['1'] != 0
        # Perform the division only where the mask is True
        sheddable_pu.loc[mask, 'size'] = sheddable_pu.loc[mask, 'DSM_shed'] / sheddable_pu.loc[mask, '1']
        # Set 'size' to 0 where the mask is False
        sheddable_pu.loc[~mask, 'size'] = 0

        ## reshape and write to .csv
        sheddable_table = (
            sheddable_pu.pivot_table(values='DSM_shed', index=['datetime', 'pattern'], columns='NAME')
            .reset_index()
            .drop(columns=['datetime'])
        )
        sheddable_table.to_csv(self.save_path[portfolio] + '/DSM_shed.csv', index=False)

    
        
        if all_regs==None:
            reg_filler = pd.Series(data = np.ones(len(demand_all.region.unique())), index = np.sort(demand_all.region.unique()))
        else:
            reg_filler = pd.Series(data = np.ones(len(all_regs)), index = np.sort(all_regs))
        
        
        """ read in DSM index elements that should not be aggregated and rename for merging where appropriate """
        di = pd.read_excel(index_path, sheet_name = indexsheet)
        di = di[pd.notnull(di['Sector.Subsector'])]
        di = di[["Sector.Subsector", "Sheddability", "aggregate_type", "header", "approx_unit", "max_scale", "max_shift"]]
        di.columns = ["end_use","Sheddability", "aggregate_type", "header", "approx_unit", "max_scale", "max_shift"]
        di_aggregate = di.groupby(["aggregate_type", "header"]).max().reset_index().drop(columns = "end_use")
        
        """ deep copy input variable (to avoid modification of source frame), scale to MW and generate scaled DSM value """
        df = raw_load.copy(deep = True)
        """replace index info in raw dataframe to ensure info from selected index is used """
        #df.columns
        df = df.drop(columns = ["Sheddability", "aggregate_type", "header"])
        df = pd.merge(df, di, how = "left")
        df = df.astype({'value':'float32', 'Sheddability':'float32'})
        df.loc[:,'value'] = df.value*1000*loss_factor
        # multiply each demand type by sheddability as per body of DSM sheets in excel version
        df.loc[:,"DSM"] = df.value * df.Sheddability

        idx_cols = ["Month", "Day", "Period", "datetime"]
        pattern_cols = ["Pattern" "datetime"]
        mdh_cols = ["Month", "Day", "Period"]
        
        """ create aggregated frame by aggregate type, header and region for use in multiple sections below
        as well as maxes frame for each aggregated end use """
        #create header aggregated frame 
        aggregated_DSM = (df.groupby(idx_cols + ["region", "aggregate_type", "header"])["DSM"].sum().unstack("region")*reg_filler).fillna(0).rename_axis(
        "region",axis=1).stack("region").rename("DSM").reorder_levels(idx_cols + ["region", "aggregate_type", "header"]).reset_index()

        # create header maxes frame for bids
        aggregated_DSM_maxes = aggregated_DSM.groupby(["region", "aggregate_type", "header"])[["DSM"]].max().reset_index()
        
        
        """ create total demands and write to .csv """
        totals = df.groupby(idx_cols + [ "region"])["value"].sum()
        totals_table = (totals.unstack("region")*reg_filler).fillna(0).reset_index()
        totals = totals.reset_index()  ### we could probably just use the pivot table approach for consistency
        totals_table.drop(columns = ["datetime"]).to_csv(save_path + "total_load.csv", index = False)
        
        """ create whole region, by end-use demand for checks (not used in model) """
        end_use_demands = df.groupby(idx_cols + [ "end_use"])[["value"]].sum().reset_index()
        end_use_table = end_use_demands.pivot_table(values = "value", index = idx_cols, columns = "end_use").reset_index()
        end_use_table.to_csv(save_path + "end_use_load_check.csv", index = False)
    
        """ create "native" load from totals minus shiftable demands """
        ## subset shiftable loads from aggregated frame
        shiftable = aggregated_DSM[aggregated_DSM.aggregate_type.isin(["shift_load"])]
        # aggregate for subtraction native load
        shiftable_totals = shiftable.groupby(idx_cols + [ "region"])["DSM"].sum().reset_index()

        ## merge with totals for subtraction and subtract
        native = pd.merge(totals, shiftable_totals, how = "left")
        native.loc[:,'value']  = native.value - native.DSM
        ## reshape and write to .csv
        native_table = (native.pivot_table(values = "value", index = idx_cols , columns = "region")*reg_filler).fillna(0).reset_index().drop(columns = ["datetime"])
        native_table.to_csv(save_path + "native_load.csv", index = False)
        """ create shiftable loads: load profile, daily sums, bid limits, max and min shifts, and annual limit for Aluminium """
        """ load profile - based on shiftable frame (already aggregated to shift category and region), create header """
        shiftable.loc[:,'NAME'] = shiftable.region + "_" + shiftable.header
        #np.unique(shiftable.header)
        ### reshape and write to .csv
        shiftable_table = shiftable.pivot_table(values = "DSM", index = idx_cols, columns = "NAME").reset_index().round(2)
        shiftable_table.drop(columns = ["datetime"]).to_csv(save_path + "DSM_shift.csv", index = False)
        """ daily sums """
        shift_sums = shiftable_table.set_index("datetime").drop(columns = mdh_cols)
        shift_sums = shift_sums.resample('D').sum()/1000 ## converted from MWh to GWh 
        round_places=5
        make_pattern_index(shift_sums).to_csv(save_path + "DSM_dayLimits.csv")
        """ bids """ #- merge in index
        bids = pd.merge(shiftable, di, how = "left")
        ## scale DSM value by max scale
        bids.loc[:,'value'] = bids.DSM*bids.max_scale
        ## find max value by region/shift category
        bids = bids.groupby(["NAME"])[["value"]].max().reset_index()
        ## add pattern column
        bids.loc[:,'pattern'] = "M1-12"
        ## set column names and write to .csv
        bids = bids[["NAME", "pattern", "value"]]
        bids.columns = ["NAME", "pattern", "1"]
        bids.to_csv(save_path + "DSM_bidQuantities.csv", index = False)
        """ max and min shift - tried changing this to monthly but produced infeasibilities
        agreed that annual max basis may be more reasonable anyway as demand response load is more coincident than unmanaged load
        so annual max basis may already be conservative """
        #np.unique(minmaxes.header)
        minmaxes = pd.merge(shiftable, di_aggregate, how = "left")
        minmaxes.loc[:,'1'] = minmaxes.DSM*minmaxes.max_shift
        ## drop null values
        minmaxes = add_time_separators(minmaxes[pd.notnull(minmaxes['1'])])
        minmaxes.loc[:,"NAME"] = minmaxes.region + "_MaxShift" + minmaxes.max_shift.astype(int).astype(str) + "h"
        ## find maxima of max shift values
        minmaxes = (minmaxes.groupby(["NAME"])["1"].max()*1.1).reset_index()
        #chkframe = minmaxes.groupby(["NAME"])["DSM"].max().reset_index()
        minmaxes.loc[:,'pattern'] = "M1-12"
        mins_frame = minmaxes.copy(deep = True)
        mins_frame.loc[:,'1'] = mins_frame['1']*-1
        mins_frame.NAME = mins_frame.NAME.str.replace("Max", "Min", case = True)
        minmaxes.append(mins_frame)[["NAME", "pattern", "1"]].to_csv(save_path + "DSM_MaxShift.csv", index = False)
        
        """ Aluminium annual limit """
        if hasAl == True:
            alframe = shiftable[shiftable.header == "Al"]
            alframe = alframe.groupby(["NAME"])[["DSM"]].sum().reset_index()
            alframe["1"] = alframe.DSM / 1000
            alframe["pattern"] = "M1-12"
            alframe[["NAME", "pattern", "1"]].to_csv(save_path + "Al_AnnualLim.csv", index = False)  
        
        """ create shed loads: units, max capacity per type, and rating per type and unit """
        """ units """
        # start with end use aggregates
        sheddable = aggregated_DSM[aggregated_DSM.aggregate_type.isin(["shed_load"])]
        # sheddable = sheddable[sheddable.region == "NER"]
        # sheddable = sheddable[sheddable.header == "Shed1h"]
        ## add unit size info
        sheddable = pd.merge(sheddable, di_aggregate, how = "left")
        sheddable.loc[:,"NAME"] = sheddable.region + "_" + sheddable.header
        # get max values and calculate number of units
        units = sheddable.groupby(["NAME"]).max().reset_index()
        units.loc[:,'1'] = round(units.DSM / units.approx_unit)
        # replace zeros with 1 if capacity is > 0
        units.at[(units["1"] == 0) & (units["DSM"] > 0) , "1"] = 1
        #units['1']
        units.loc[:,"pattern"] = "M1-12"
        units[["NAME", "pattern", "1"]].to_csv(save_path + "shed_units.csv", index = False)
        """ max capacity per type - calculate from DSM and units """
        units.loc[:,"size"] = units.DSM / units['1']
        unit_size = units[["NAME", "pattern", "size"]]
        unit_size.columns = ["NAME", "pattern", "1"]
        unit_size.to_csv(save_path + "shed_max_cap.csv", index = False)
        
        """ rating per type and unit by dividing through aggregated timeseries by number of units """
        sheddable_pu = pd.merge(sheddable, units[["NAME", "1"]], how = "left")
        sheddable_pu.DSM = sheddable_pu.DSM / sheddable_pu["1"]
        sheddable_pu.DSM = sheddable_pu.DSM.fillna(0)
        ## reshape and write to .csv
        sheddable_table = sheddable_pu.pivot_table(values = "DSM", index = idx_cols, columns = "NAME").reset_index().drop(columns = ["datetime"])
        sheddable_table.to_csv(save_path + "DSM_shed.csv", index = False)
        
        return("Processed successfully to: {}".format(save_path))
    
    def _read_manual_demand_data(self):
        """
        Function to read in manual demand data from a csv file. This could be either national or regional demand data.
        The function will read in the data and return a pandas dataframe with the demand data.
        """


        demand_file_path = self.c.cfg['load']['load_path']
        # demand_file_path = 'Y:/Modelling/Ukraine/2023_UKR_ST_Security/03_Modelling/01_InputData/02_DemandSide/'
        timezone = self.c.cfg['load']['timezone'] ## Ukrainian standard time 
        demand_hourly = pd.read_csv(demand_file_path)
        index_cols = self.c.cfg['load']['index_cols']
        if self.c.cfg['load']['load_type'] == 'regional':
            value_cols = self.c.cfg['parameters']['regions']
        else:
            demand_hourly = demand_hourly.rename(columns={self.c.cfg['load']['load_data_col']:'Value'})
            value_cols = ['Value']


        demand_hourly = demand_hourly[index_cols + value_cols]

        ### Create datetime index for hourly demand 
        demand_hourly.index = pd.to_datetime(demand_hourly[index_cols])
        demand_hourly = demand_hourly[value_cols]

        ## Filter out demand data
        # clean_timeseries_methods = self.c.cfg['load']['clean_ts_methods']
        clean_timeseries_methods = ['positive', 'sudden_change', 'std']
        
        demand_hourly = clean_timeseries(demand_hourly, column=value_cols, 
                                         methods=clean_timeseries_methods
        )
        demand_hourly = demand_hourly.shift(timezone).dropna()
        load_res = self.c.cfg['load']['load_res']

        try:
            load_start = self.c.cfg['load']['date_start']
            load_end = self.c.cfg['load']['date_end']
            demand_idx = pd.date_range(start=load_start, end=load_end, freq=load_res)
        except KeyError:
            # log.warning("No start/end dates specified in config file. Using data from file.")
            demand_idx = pd.date_range(start=demand_hourly.index[0], end=demand_hourly.index[-1], freq=load_res)


        ### Create a new, continuous dataframe for dealing with gaps in data (for interpolation)
        demand_idx = pd.date_range(start=load_start, end=load_end, freq=load_res)
        demand_hourly = demand_hourly.reindex(demand_idx)
        demand_hourly = demand_hourly.interpolate('linear')

        return demand_hourly
    



    

#### General functions
#############

def make_mdh_index(df):
    if df.index.freqstr == 'D':
        df['Month'] = df.index.month
        df['Day'] = df.index.day
        df = df.set_index(['Month', 'Day'])
    else:
        df['Month'] = df.index.month
        df['Day'] = df.index.day
        df['Period'] = df.index.hour + 1
        df = df.set_index(['Month', 'Day', 'Period'])

    return df

def make_pattern_index(df):
    if df.index.freqstr == 'D':
        df['Pattern'] = df.index.to_series().apply(lambda x: 'M{},D{}'.format(x.month,x.day))
    else:
        df['Pattern'] = df.index.to_series().apply(lambda x: 'M{},D{},H{}'.format(x.month,x.day,x.hour+1))
        
    df = df.set_index('Pattern')

    return df

def read_end_use_demand_WEO_format(file_path, sheet_vector, indexsheet  = "DSM_Index", StartRow = 1, RegionSplit = "", RegionVector = [], Scale_factor=1,
                                   end_use_adj_sheet = "", end_use_col = ""):
        
        ## create empty dataframe to contain results
        df = pd.DataFrame()
        
        ## read in region sheets one by one, melt and append
        ## very slow due to slowness of pd.read_excel
        for i in sheet_vector:
            # read in region sheet including end use headings and removing unwanted row
            demand_reg = pd.read_excel(file_path, sheet_name=i,header=[1]).iloc[1:,:].reset_index()
            #demand_reg = pd.read_excel(file_path, sheet_name=i,skiprows =StartRow-2, header=[1]).iloc[1:,:].reset_index()
            #demand_reg = pd.read_excel(file_path, sheet_name=i,startrow =(StartRow)).iloc[1:,:]
            # convert to long format
            dcm = pd.melt(demand_reg, id_vars = "Unnamed: 0")
            dcm.columns = (["datetime", "end_use", "value"])
            # add region column
            dcm.loc[:,'region'] = i
            # append to result dataframe
            df = df.append(dcm)
            
            
        ## read DSM index in order to trim end uses
        di = pd.read_excel(file_path, sheet_name = indexsheet)
        di = di[pd.notnull(di['Sector.Subsector'])]
        di = di[["Sector.Subsector", "Sheddability", "aggregate_type", "header"]]
        di.columns = ["end_use", "Sheddability", "aggregate_type", "header"]
        
        ## drop entries not corresponding with single end uses
        df = df[df.end_use.isin(np.unique(di.end_use))]
        init_sum = df["value"].sum()  
        
        ## merge
        df = pd.merge(df, di, how = "left")
        
        
        ## add datetime column for ordering, uses leap year to ensure feb is preserved if present
        ## this is a bit annoying but used to force [Month,Day,Hour] index to come out in the right order after pivoting
        colnames = df.columns.tolist()
        for i in ["Period","Day","Month"]:
            colnames.insert(0, i)
        df["Month"] = df.datetime.dt.month
        df["Day"] = df.datetime.dt.day
        df["Period"] = df.datetime.dt.hour + 1

        df = df[colnames]
        
        ## end use scaling if sheet / column are defined
        if len(end_use_adj_sheet) > 0:
            if end_use_col =="":
                print("Please define the column to refer to in the end use adjustment sheet")
            # read in end use totals 
            eua = pd.read_excel(file_path, sheet_name = end_use_adj_sheet).dropna()
            eua.loc[:,"eut"] = eua[end_use_col]
            print("End use scaling checks: Sum of end use target values: ", str(round(eua.eut.sum(), 0)))
            # merge with hourly frame
            dfeut = df.merge(eua[["end_use", "eut"]])
            # add column of sums by end use for comparison with target totals
            dfeut.loc[:,"end_use_sum"] = dfeut.groupby(["end_use"])["value"].transform("sum")/1000
            #adjust value based on ratio
            dfeut.loc[:,"value"] = dfeut["value"] * (dfeut.eut / dfeut.end_use_sum)
            print("Sum of values before scale: ", str(round(df["value"].sum()/1000, 0)))
            df = dfeut[colnames]
            print("Sum of values after scale: ", str(round(df["value"].sum()/1000, 0)))
        
        region_split_sum = 0
        
        if len(RegionSplit)>0:
            df.loc[:, 'scenario'] = df['region']
            df = df.drop(columns = 'region')
            regionsheet = pd.read_excel(file_path, sheet_name = RegionSplit)
            rs = pd.melt(regionsheet, id_vars = "end_use", value_vars = RegionVector, var_name= "region").rename(columns = {"value": "regionsplit"})
            df = pd.merge(df, rs, how = "left")
            df.loc[:,"orig_val"] = df["value"]
            df.loc[:,"value"] = df["orig_val"]*df["regionsplit"]
            split_sum = df["value"].sum()
        
        df["before_scale"] = df["value"]
        df.loc[:,"value"] = df["before_scale"]*Scale_factor
        scaled_sum = df["value"].sum()
        df.loc[:,"value"] = df.value.astype(np.float32)
        
        # return combined
        return df

def add_time_separators(inputFrame, datetimeCol = "datetime", pattern_date = False, set_year = False, set_month = False,
                    timeconvention = "time_start"):

    df = inputFrame.copy(deep = True)
    dtcol = datetimeCol[:]
    if pattern_date == True:
        # create dummy datetime sequence using a leap or specified year and add merge info
        sd = dt(year=2020, month=1, day=1)
        if set_year != False:
            sd = dt(year = set_year, month = 1, day = 1)
            if set_month != False:
                sd = dt(year = set_year, set_month = 1, day = 1)
        dtdf = pd.DataFrame(pd.date_range(start=sd, end=sd+pd.offsets.DateOffset(years=1)+pd.offsets.DateOffset(hours=-1), freq='h'))
        dtdf.columns = ["datetime"]
        dtdf['month'] = pd.DatetimeIndex(dtdf.datetime).month
        dtdf['mday'] = pd.DatetimeIndex(dtdf.datetime).day
        dtdf['hour']= pd.DatetimeIndex(dtdf.datetime).hour
        # derive month day and hour info from pattern index
        df['month'] = df[dtcol].str.split(",").str[0].str.replace("M", "").astype(float)
        df['mday'] = df[dtcol].str.split(",").str[1].str.replace("D", "").astype(float)
        df['hour'] = df[dtcol].str.split(",").str[2].str.replace("H", "").astype(float) - 1
        
        #merge in dummy date sequence to allow remaining separators to be added as normal
        df = pd.merge(df, dtdf, how = "left")
        dtcol = "datetime"
        
    #df = df.set_index("datetime")
    if timeconvention == "time_end":
        df["original_datetime"] = df[dtcol]
        df["datetime"]= df["datetime"] - timedelta(hours=1)

    df['year'] = pd.DatetimeIndex(df[dtcol]).year
    df['month'] = pd.DatetimeIndex(df[dtcol]).month
    df['montht'] = pd.DatetimeIndex(df[dtcol]).month_name()
    df['week'] = df[dtcol].dt.isocalendar().week
    df['mday'] = pd.DatetimeIndex(df[dtcol]).day
    df['day'] = pd.DatetimeIndex(df[dtcol]).day
    df['yday'] = pd.DatetimeIndex(df[dtcol]).dayofyear
    df['hour']= pd.DatetimeIndex(df[dtcol]).hour
    df['pattern'] = "M" + df.month.astype(str) + ",D" + df.mday.astype(str) + ",H" +(df.hour+1).astype(str)
    df['wday_num'] = pd.DatetimeIndex(df[dtcol]).dayofweek
    df['wdaytype'] = "blank"
    df.loc[df['wday_num'].isin([0,1,2,3,4]), "wdaytype"] = "Weekday"
    df.loc[df['wday_num'].isin([5]), "wdaytype"] = "Saturday"
    df.loc[df['wday_num'].isin([6]), "wdaytype"] = "Sunday"
    df["seasonNH"] = "blank"
    df["seasonSH"] = "blank"
    df["two_seasonNH"] = "blank"
    df.loc[df['month'].isin([12, 1, 2]), "seasonNH"] = "Winter"
    df.loc[df['month'].isin([12, 1, 2]), "seasonSH"] = "Summer"
    df.loc[df['month'].isin([3,4,5]), "seasonNH"] = "Spring"
    df.loc[df['month'].isin([3,4,5]), "seasonSH"] = "Autumn"
    df.loc[df['month'].isin([6,7,8]), "seasonNH"] = "Summer"
    df.loc[df['month'].isin([6,7,8]), "seasonSH"] = "Winter"
    df.loc[df['month'].isin([9, 10, 11]), "seasonNH"] = "Autumn"
    df.loc[df['month'].isin([9, 10, 11]), "seasonSH"] = "Spring"
    
    df.loc[df['month'].isin([1, 2,3,10,11,12]), "two_seasonNH"] = "Winter"
    df.loc[df['month'].isin([4, 5, 6, 7, 8, 9]), "two_seasonNH"] = "Summer"
    

    #chk = df[df.month.isna()]

        
    if pattern_date == True:
        # strip off year (datetime column and year column) unless was specified
        if set_year == False:
            df = df.drop(["datetime", "year"], axis = 1)
    
    
    return(df)

def convert_raw_load_to_PLEXOS_inputs(raw_load, index_path, save_path, indexsheet = "DSM_Index", hasAl = True, all_regs=None, loss_factor = 1):